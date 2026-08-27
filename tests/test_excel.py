"""Tests for XLSX generation."""

from __future__ import annotations

from openpyxl import load_workbook

from app.schemas.extraction import (
    DocumentResult,
    EntityFields,
    ExtractedValue,
    TableBlock,
    TextBlock,
)
from app.services.excel_service import generate_xlsx


def test_generate_xlsx_creates_sheet_per_table(tmp_path):
    result = DocumentResult(
        document_type="invoice",
        language="ru",
        pages=1,
        text_blocks=[
            TextBlock(
                type="table",
                table=TableBlock(title="Items", headers=["No", "Item", "Price"], rows=[["1", "Pen", "2.50"]]),
                source_page=1,
            ),
            TextBlock(
                type="table",
                table=TableBlock(title="Totals", headers=["Label", "Amount"], rows=[["Total", "2.50"]]),
                source_page=1,
            ),
        ],
        entities=EntityFields(names=[ExtractedValue(value="John Doe", confidence=0.9, source_page=1)]),
    )
    output_path = tmp_path / "output.xlsx"

    generate_xlsx(result, output_path=str(output_path))

    assert output_path.exists()
    wb = load_workbook(str(output_path))
    assert "Summary" in wb.sheetnames
    assert "Items" in wb.sheetnames
    assert "Totals" in wb.sheetnames

    items_sheet = wb["Items"]
    assert items_sheet["A1"].value == "No"
    assert items_sheet["B2"].value == "Pen"

    summary_sheet = wb["Summary"]
    values = [row[1].value for row in summary_sheet.iter_rows(min_row=2)]
    assert "John Doe" in values


def test_generate_xlsx_falls_back_to_text_sheet_when_no_tables(tmp_path):
    result = DocumentResult(
        document_type="letter",
        language="en",
        pages=1,
        text_blocks=[TextBlock(type="paragraph", text="Just a plain paragraph.", source_page=1)],
    )
    output_path = tmp_path / "no_tables.xlsx"

    generate_xlsx(result, output_path=str(output_path))

    wb = load_workbook(str(output_path))
    assert "Extracted Text" in wb.sheetnames
    sheet = wb["Extracted Text"]
    assert sheet["B2"].value == "Just a plain paragraph."


def test_generate_xlsx_duplicate_table_titles_get_unique_sheet_names(tmp_path):
    result = DocumentResult(
        document_type="report",
        language="en",
        pages=2,
        text_blocks=[
            TextBlock(type="table", table=TableBlock(title="Data", headers=["A"], rows=[["1"]]), source_page=1),
            TextBlock(type="table", table=TableBlock(title="Data", headers=["A"], rows=[["2"]]), source_page=2),
        ],
    )
    output_path = tmp_path / "dupes.xlsx"

    generate_xlsx(result, output_path=str(output_path))

    wb = load_workbook(str(output_path))
    data_sheets = [name for name in wb.sheetnames if name.startswith("Data")]
    assert len(data_sheets) == 2
