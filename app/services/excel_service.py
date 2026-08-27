"""XLSX generation from a validated DocumentResult, using openpyxl.

Each detected table becomes its own sheet. If the document has no tables,
a fallback "Extracted Text" sheet is generated from paragraphs so the user
still gets a usable Excel export. A "Summary" sheet with entities is always
included when any entities were detected.
"""

from __future__ import annotations

import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas.extraction import DocumentResult, EntityFields

_HEADER_FILL = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")
_HEADER_FONT = Font(bold=True)
_MAX_SHEET_NAME = 31
_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def _safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = _INVALID_SHEET_CHARS.sub("_", name).strip() or "Sheet"
    cleaned = cleaned[:_MAX_SHEET_NAME]
    candidate = cleaned
    suffix = 1
    while candidate.lower() in used:
        suffix_str = f" ({suffix})"
        candidate = cleaned[: _MAX_SHEET_NAME - len(suffix_str)] + suffix_str
        suffix += 1
    used.add(candidate.lower())
    return candidate


def _autosize_and_format(ws: Worksheet, num_columns: int, has_header: bool = True) -> None:
    if has_header:
        for col in range(1, num_columns + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
        ws.freeze_panes = "A2"
        if ws.max_row > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(num_columns)}{ws.max_row}"

    for col in range(1, num_columns + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)


def _write_table_sheet(wb: Workbook, sheet_name: str, headers: list[str], rows: list[list[str]]) -> None:
    ws = wb.create_sheet(title=sheet_name)
    columns = len(headers) or (len(rows[0]) if rows else 1)
    if headers:
        ws.append(headers)
    for row in rows:
        padded = list(row) + [""] * max(0, columns - len(row))
        ws.append(padded[:columns])
    _autosize_and_format(ws, columns, has_header=bool(headers))


def _write_entities_sheet(wb: Workbook, entities: EntityFields) -> None:
    ws = wb.create_sheet(title="Summary")
    ws.append(["Category", "Value", "Confidence", "Page", "Uncertain"])
    rows_written = 0
    for category, values in (
        ("Name", entities.names),
        ("Date", entities.dates),
        ("Document Number", entities.document_numbers),
        ("Amount", entities.amounts),
        ("Address", entities.addresses),
        ("Signature", entities.signatures),
        ("Stamp", entities.stamps),
    ):
        for item in values:
            ws.append(
                [
                    category,
                    item.value,
                    round(item.confidence, 2),
                    item.source_page or "",
                    "Yes" if item.uncertain else "No",
                ]
            )
            rows_written += 1
    if rows_written == 0:
        ws.append(["No entities detected", "", "", "", ""])
    _autosize_and_format(ws, 5, has_header=True)


def generate_xlsx(result: DocumentResult, output_path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    used_names: set[str] = set()

    has_entities = any(
        len(getattr(result.entities, field))
        for field in (
            "names",
            "dates",
            "document_numbers",
            "amounts",
            "addresses",
            "signatures",
            "stamps",
        )
    )
    if has_entities:
        _write_entities_sheet(wb, result.entities)
        used_names.add("summary")

    if result.tables:
        for index, table in enumerate(result.tables, start=1):
            base_name = table.title or f"Table {index}"
            sheet_name = _safe_sheet_name(base_name, used_names)
            _write_table_sheet(wb, sheet_name, table.headers, table.rows)
    else:
        sheet_name = _safe_sheet_name("Extracted Text", used_names)
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["Page", "Text"])
        for paragraph in result.paragraphs:
            ws.append([paragraph.source_page or "", paragraph.text])
        for heading in result.headings:
            ws.append([heading.source_page or "", heading.text])
        _autosize_and_format(ws, 2, has_header=True)

    if not wb.sheetnames:
        wb.create_sheet(title="Extracted Data")

    wb.save(output_path)
